# -*- coding: utf-8 -*-
"""
Created on Mon Jun  3 14:16:05 2019

@author: Shaoqi WU
"""
import os
from win32com.client import Dispatch

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

import xlrd
import openpyxl


def read_coord(sheet_name): 
    profil_width = 500
    style = openpyxl.styles.Font(bold=True)
    row0 = ["Name", "COORx", "COORy", "COORz", "Vector"]
    for i in range(0, len(row0)):
        sheet_i.cell(1, 1+i, row0[i])
        sheet_i.cell(1, 1+i).font = style
        
    if sheet_name == "SPF" or "SPR" or "SRE":
        point_number = 12
    if sheet_name == "SKF":
        point_number = 6
    if sheet_name == "SRI":
        point_number = 10
        
    for i in range(point_number):
        name = sheet_name + "1_"+str(i+1)
        sheet_i.cell(i+2, 1, name)
        
        mesure_x = "Wing\\" + sheet_name + "1_"+str(i+1)+"\Gx"
        param_x = product1.Parameters.GetItem(mesure_x).Value
        sheet_i.cell(i+2, 2, param_x)
        
        mesure_y = "Wing\\" + sheet_name + "1_"+str(i+1)+"\Gy"
        param_y = product1.Parameters.GetItem(mesure_y).Value
        sheet_i.cell(i+2, 3, param_y)
        
        mesure_z = "Wing\\" + sheet_name + "1_"+str(i+1)+"\Gz"
        param_z = product1.Parameters.GetItem(mesure_z).Value
        sheet_i.cell(i+2, 4, param_z)
        
        sheet_i.cell(i+2, 5, (param_x**2 + param_y**2)**0.5)
        
        name = sheet_name + "2_"+str(i+1)
        sheet_i.cell(i+point_number+2, 1, name)
        sheet_i.cell(i+point_number+2, 2, param_x)
        sheet_i.cell(i+point_number+2, 3, param_y)
        sheet_i.cell(i+point_number+2, 4, param_z - profil_width)
        sheet_i.cell(i+point_number+2, 5, (param_x**2 + param_y**2)**0.5)
        
    directory = "Z:\\shaoqi-WU\\Shaoqi_Stage\\modele\\wing_Shaoqi\\CAO_Aile_v2_03062019"
    excel.save(directory+"\\coord_points.xlsx")
     

def read_from_excel(Matrix, excel_name, sheet_name):
    matrix = pd.read_excel(excel_name, sheet_name[0], index_col = 0)
    array_ = matrix.values
    
#connecting to windows com
CATIA = Dispatch('CATIA.Application')
CATIA.Visible = False

#open catia
file_path = "Z:\shaoqi-WU\Shaoqi_Stage\modele\wing_Shaoqi\CAO_Aile_v2_03062019\Wing.CATProduct"
product_document = CATIA.Documents.Open(file_path)

product1 = product_document.Product

# create excel
excel = openpyxl.Workbook()
sheet_name = ["SPF", "SPR", "SRE", "SKF", "SRI"]

for i, sheet in enumerate(sheet_name):
    sheet_i = excel.create_sheet(sheet, index=i)
    read_coord(sheet)

# read generated excel
directory = "Z:\\shaoqi-WU\\Shaoqi_Stage\\modele\\wing_Shaoqi\\CAO_Aile_v2_03062019"
excel_name = directory+"\\coord_points.xlsx"




